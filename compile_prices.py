import openpyxl
import json
import os

def compile_excel_to_js():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(current_dir, "especificacao_precos.xlsx")
    js_path = os.path.join(current_dir, "precos.js")
    
    if not os.path.exists(excel_path):
        print(f"Erro: O arquivo {excel_path} não foi encontrado!")
        return
        
    print(f"Carregando {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    database = {
        "catalog": {},
        "compositions": {},
        "sensors": {},
        "clpRules": [],
        "ihmMapping": {},
        "supervisorioRules": []
    }
    
    # 1. Catalog
    ws_cat = wb["Catálogo de Componentes"]
    for row in range(2, ws_cat.max_row + 1):
        code = ws_cat.cell(row=row, column=1).value
        desc = ws_cat.cell(row=row, column=2).value
        brand = ws_cat.cell(row=row, column=3).value
        unit = ws_cat.cell(row=row, column=4).value
        price = ws_cat.cell(row=row, column=5).value
        
        if code:
            database["catalog"][str(code).strip()] = {
                "desc": str(desc).strip() if desc else "",
                "brand": str(brand).strip() if brand else "",
                "unit": str(unit).strip() if unit else "un",
                "price": float(price) if price is not None else 0.0
            }
            
    # Inject mandatory items if missing
    if "SINALEIRO-BRANCO" not in database["catalog"]:
        database["catalog"]["SINALEIRO-BRANCO"] = {
            "desc": "SINALEIRO LED MONOBLOCO BRANCO 220V/24V (INDICADOR LIGADO)",
            "brand": "WEG",
            "unit": "un",
            "price": 25.0
        }
        
    # Inject Heavy Line items (NBR 5597 / NBR 5598)
    heavy_eletroduto_prices = {
        '1/2': 30.0,
        '3/4': 36.0,
        '1': 48.0,
        '1.1/4': 66.0,
        '1.1/2': 84.0,
        '2': 114.0
    }
    heavy_condulete_prices = {
        '1/2': 37.5,
        '3/4': 45.0,
        '1': 62.5,
        '1.1/4': 87.5,
        '1.1/2': 112.5,
        '2': 162.5
    }
    for size in ['1/2', '3/4', '1', '1.1/4', '1.1/2', '2']:
        database["catalog"][f"ELETRODUTO-PESADO-{size}"] = {
            "desc": f"ELETRODUTO GALVANIZADO PESADO {size} (NBR 5597/5598)",
            "brand": "WETZEL",
            "unit": "m",
            "price": heavy_eletroduto_prices[size]
        }
        database["catalog"][f"CONDULETE-PESADO-T-{size}"] = {
            "desc": f"CONDULETE PESADO TIPO T {size} COM TAMPA (NBR 5597/5598)",
            "brand": "WETZEL",
            "unit": "un",
            "price": heavy_condulete_prices[size]
        }
        database["catalog"][f"CONDULETE-PESADO-LR-{size}"] = {
            "desc": f"CONDULETE PESADO TIPO LR {size} COM TAMPA (NBR 5597/5598)",
            "brand": "WETZEL",
            "unit": "un",
            "price": heavy_condulete_prices[size]
        }
        database["catalog"][f"CONDULETE-PESADO-E-{size}"] = {
            "desc": f"CONDULETE PESADO TIPO E {size} COM TAMPA (NBR 5597/5598)",
            "brand": "WETZEL",
            "unit": "un",
            "price": heavy_condulete_prices[size]
        }
        
    # Inject Network Cable CAT 6
    database["catalog"]["CABO-REDE-CAT6"] = {
        "desc": "CABO DE REDE ETHERNET CAT-06 (AZUL)",
        "brand": "FURUKAWA",
        "unit": "m",
        "price": 5.0
    }
    
    # Inject Transformer 400VA 440/220V
    database["catalog"]["TRANSFORMADOR-440-220-400VA"] = {
        "desc": "TRANSFORMADOR MONOFÁSICO 440/220V - 400VA",
        "brand": "WEG",
        "unit": "un",
        "price": 600.0
    }

    # Inject Trilho DIN Brum 1m
    database["catalog"]["TRILHO-DIN-1M"] = {
        "desc": "TRILHO DIN PERFURADO BARRA DE 1M MARCA BRUM",
        "brand": "BRUM",
        "unit": "un",
        "price": 15.0
    }

    # Remove old CABO-1.0 if present to enforce unique models only
    database["catalog"].pop("CABO-1.0", None)
            
    # 2. Compositions
    ws_comp = wb["Composições de Partida"]
    for row in range(2, ws_comp.max_row + 1):
        p_type = ws_comp.cell(row=row, column=1).value
        power = ws_comp.cell(row=row, column=2).value
        voltage = ws_comp.cell(row=row, column=3).value
        comp_code = ws_comp.cell(row=row, column=4).value
        qty = ws_comp.cell(row=row, column=5).value
        
        if p_type and power and voltage and comp_code:
            key = f"{str(p_type).strip()}_{str(power).strip()}_{str(voltage).strip()}"
            if key not in database["compositions"]:
                database["compositions"][key] = []
            database["compositions"][key].append({
                "code": str(comp_code).strip(),
                "qty": float(qty) if qty is not None else 1.0
            })
            
    # 3. Sensors
    ws_sens = wb["Sensores e Leituras"]
    for row in range(2, ws_sens.max_row + 1):
        equip = ws_sens.cell(row=row, column=1).value
        sensor = ws_sens.cell(row=row, column=2).value
        comp_code = ws_sens.cell(row=row, column=3).value
        qty = ws_sens.cell(row=row, column=4).value
        
        if equip and sensor and comp_code:
            key = f"{str(equip).strip()}_{str(sensor).strip()}"
            if key not in database["sensors"]:
                database["sensors"][key] = []
            database["sensors"][key].append({
                "code": str(comp_code).strip(),
                "qty": float(qty) if qty is not None else 1.0
            })
            
    # 4. CLP Rules
    ws_clp = wb["CLP e Conectores"]
    for row in range(2, ws_clp.max_row + 1):
        equip = ws_clp.cell(row=row, column=1).value
        min_qty = ws_clp.cell(row=row, column=2).value
        max_qty = ws_clp.cell(row=row, column=3).value
        clp_code = ws_clp.cell(row=row, column=4).value
        clp_qty = ws_clp.cell(row=row, column=5).value
        conn_code = ws_clp.cell(row=row, column=6).value
        conn_qty = ws_clp.cell(row=row, column=7).value
        
        if equip:
            database["clpRules"].append({
                "equipType": str(equip).strip(),
                "minQty": int(min_qty) if min_qty is not None else 0,
                "maxQty": int(max_qty) if max_qty is not None else 999,
                "clpCode": str(clp_code).strip() if clp_code else None,
                "clpQty": float(clp_qty) if clp_qty is not None else 0.0,
                "connCode": str(conn_code).strip() if conn_code else None,
                "connQty": float(conn_qty) if conn_qty is not None else 0.0
            })
            
    # 5. IHMs
    ws_ihm = wb["IHMs e Supervisório"]
    for row in range(2, ws_ihm.max_row + 1):
        opt = ws_ihm.cell(row=row, column=1).value
        comp_code = ws_ihm.cell(row=row, column=2).value
        qty = ws_ihm.cell(row=row, column=3).value
        
        if opt and comp_code:
            opt_str = str(opt).strip()
            database["ihmMapping"][opt_str] = {
                "code": str(comp_code).strip(),
                "qty": float(qty) if qty is not None else 1.0
            }
            
    # 6. Supervisório Rules (New!)
    ws_sup = wb["Modelos de Supervisório"]
    for row in range(2, ws_sup.max_row + 1):
        min_qty = ws_sup.cell(row=row, column=1).value
        max_qty = ws_sup.cell(row=row, column=2).value
        comp_code = ws_sup.cell(row=row, column=3).value
        qty = ws_sup.cell(row=row, column=4).value
        
        if min_qty is not None and comp_code:
            database["supervisorioRules"].append({
                "minQty": int(min_qty),
                "maxQty": int(max_qty) if max_qty is not None else 999,
                "code": str(comp_code).strip(),
                "qty": float(qty) if qty is not None else 1.0
            })
                
    # Write to precos.js
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write("/* AUTO-GENERATED PRICING DATABASE FROM EXCEL */\n")
        f.write("const PRECOS_DATABASE = ")
        json.dump(database, f, indent=2, ensure_ascii=False)
        f.write(";\n")
        
    print(f"precos.js gerado com sucesso em: {js_path}")

if __name__ == "__main__":
    compile_excel_to_js()
